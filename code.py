from string import Template

import fitz
import pdfkit
from openpyxl import load_workbook


def sheet(filename="sample.xlsx"):
    wb = load_workbook(filename, data_only=True)
    sheet = wb.active
    return sheet


def headers(sheet):
    # if we want to skipe last cols if cell.value is not None
    header_cells = sheet[1]
    header = [cell.value for cell in header_cells]
    return header


def employee_data(employee):
    k = [
        "row",  # سطر
        "name",  # اسم کارمند
        "days_worked",  # روز کارد
        "daily_pay",  #  دستمزد روزانه
        "extra_wh",  #  ساعت اضافه کار
        "daily_mp",  #  نرخ ماموریت روزانه
        "extra_whp",  #  نرخ اضافه کار
        "base_pay",  #  پایه حقوق
        "work_pay",  #  مبلغ کارکرد
        "extra_work_pay",  #  مبلغ اضافه کار
        "house_right",  #  حق مسکن
        "extra_help",  #  کمک هزینه اقلام مصرفی
        "child_right",  #   حق بچه
        "commiute",  #  ایاب ذهاب
        "extra_undirect",  #  مزایای غیر مستقیم
        "work_extra",  #  فوق العاده شغل
        "other_benefit",  # سایر مزایا
        "u_payment",  #  حقوق ناخالص
        "seven_ensure",  #  هفت درصد بیمه
        "to_tax",  #  مشمول مالیات
        "tax",  #  مالیات
        "payemt_extra",  # خالص حقوق و مزایا
        "rem",  # علی الحساب
        "loan",  #  وام قرض الحسنه
        "other_insur",  #  سایر بیمه تکمیلی
        "other_req",  # سایر مطالبات
        "mission",  #  ماموریت
        "other",  #  سایر
        "payment",  #  قابل پرداخت
        "name2",  #  اسم
    ]
    v = [cell.value for cell in employee]
    data = dict(zip(k, v))
    employee = {}
    for k, v in data.items():
        if v == None:
            employee[k] = ""
        else:
            employee[k] = v
    # adding extra options
    ew = employee.get("extra_whp")
    employee["year"] = "1399"
    employee["month"] = "مهر"
    employee["extra_whp"] = f"{ew:.2f}"
    return employee


def employees(sheet):
    # skip hedaer
    rows = list(sheet.iter_rows())[2:]
    # if value is not that mean we found an employee
    employees = [row for row in rows if row[0].value is not None]
    return employees[:-1]  # skipe last row


# Create employee dict by employee name
def employee(employees, name=None):
    if name is not None:
        for emp in employees:
            if emp[1].value == name:
                return employee_data(emp)
    return None


# cols: row, name, email
def items(employees, column_name="name"):
    columns = {"row": 0, "name": 1, "email": 20}
    index = columns.get(column_name)
    return [employee[index].value for employee in employees]


def read_template(name="template.html"):
    with open("template.html", "r") as template_file:
        template = Template(template_file.read())
    return template


# create payroll for a employee (get employee dict)
def create_payroll_html(employee, template, filename="payroll.html"):
    payroll = template.substitute(employee)
    with open(filename, "w") as html_file:
        html_file.write(payroll)
    return filename


def html_to_pdf(html_filename="payroll.html", pdf_filename="payroll.pdf"):
    options = {
        "encoding": "UTF-8",
        "enable-local-file-access": "",
    }
    pdfkit.from_file(html_filename, pdf_filename, options=options)
    return pdf_filename


def pdf_to_image(pdf_path, img_name="payroll.jpg"):
    doc = fitz.open(pdf_path)
    page = doc[0]  # payroll is just one page
    pix = page.getPixmap()
    pix.writeImage(img_name)
    return pix


def get_image_bytes(pix):
    pix1 = (
        fitz.Pixmap(pix, 0) if pix.alpha else pix
    )  # PPM does not support transparency
    imgdata = pix1.getImageData("ppm")  # extremely fast!
    return imgdata


def main():
    sh = sheet()
    emps = employees(sh)
    # x = employee(emps, "john")


if __name__ == "__main__":
    sh = sheet()
    emps = employees(sh)
    # x = employee(emps, "john")
