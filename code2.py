from openpyxl import load_workbook
from string import Template
import pdfkit
import os





def sheet():
    wb = load_workbook(filename="sample.xlsx", data_only=True)
    sheet = wb.active
    return sheet


def headers(sheet):
    # if we want to skipe last cols if cell.value is not None
    header_cells = sheet[1]
    header = [cell.value for cell in header_cells]
    return header


def employees(sheet):
    # skip hedaer
    rows = list(sheet.iter_rows())[2:]
    # if value is not that mean we found an employee
    employees = [row for row in rows if row[0].value is not None]
    print(employees[0][1].value)
    return employees[:-1]  # skipe last row


# cols: row, name, email
def items(employees, column_name="name"):
    columns = {"row": 0, "name": 1, "email": 20}
    index = columns.get(column_name)
    return [employee[index].value for employee in employees]

def create_emplo_dict(sheet):
    k = [
        "row",  # سطر
        "name",  # اسم کارمند
        "days_worked",  # روز کارد
        "daily_pay",  #  دستمزد روزانه
        "extra_wh",  #  ساعت اضافه کار
        "daily_mp",  #  نرخ ماموریت روزانه
        "extra_whp",  #  نرخ اضافه کار
        "base_pay",  #  پایه حقوق
        "work_pay",  #  مبلغ کارکرد
        "extra_work_pay",  #  مبلغ اضافه کار
        "house_right",  #  حق مسکن
        "extra_help",  #  کمک هزینه اقلام مصرفی
        "child_right",  #   حق بچه
        "commiute",  #  ایاب ذهاب
        "extra_undirect",  #  مزایای غیر مستقیم
        "work_extra",  #  فوق العاده شغل
        "other_benefit",  # سایر مزایا
        "u_payment",  #  حقوق ناخالص
        "seven_ensure",  #  هفت درصد بیمه
        "to_tax",  #  مشمول مالیات
        "tax",  #  مالیات
        "payemt_extra",  # خالص حقوق و مزایا
        "rem",  # علی الحساب
        "loan",  #  وام قرض الحسنه
        "other_insur",  #  سایر بیمه تکمیلی
        "other_req",  # سایر مطالبات
        "mission",  #  ماموریت
        "other",  #  سایر
        "payment",  #  قابل پرداخت
        "name2",  #  اسم
    ]
    allcells = employees(sheet)
    employee_dict = {}
    for num, cells in enumerate(allcells):
        employee_name = allcells[num][1].value
        v = [cell.value for cell in allcells[num]]
        employee_t = dict(zip(k, v))
        employee_dict[employee_name] = employee_t
    return employee_dict


def find_right_row(row_tree, sheet):
    employee_dict = create_emplo_dict(sheet)
    names = []
    if len(row_tree) > 0:
        for item in row_tree:
            names.append(item[1])
        return  { key:value for key, value in employee_dict.items() if key in names }

def create_pdf(employee_dict, row_tree):
    print('employee_dict in create_pdf function ', employee_dict)
    if not os.path.exists('my_folder'):
        os.makedirs('my_folder')
    os.popen('copy logo.png .\\my_folder\\logo.png')
    # os.popen('copy source.txt destination.txt')
    filepathes = []
    names = employee_dict.keys()
    print('names', names)
    for name in names:
        ew = employee_dict[name].get("extra_whp")
        f = open("payroll_template.html", "r", encoding='utf-8')
        payroll_template = Template(f.read())
        payroll = payroll_template.substitute(
            employee_dict[name], year=1399, month="مهر", extra_whp=f"{ew:.2f}"
        )
        f.close()
        rownum = employee_dict[name].get("row")
        htmlfile = ".\\my_folder\\payroll_{}.html".format(rownum)
        f = open(htmlfile, "w", encoding='utf-8')
        f.write(payroll)
        f.close()

        pdfkit_options = {
            'encoding': 'UTF-8',
            "enable-local-file-access":"",
        }
        path_wkthmltopdf = b'C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe'
        config = pdfkit.configuration(wkhtmltopdf=path_wkthmltopdf)
        print(employee_dict[name])
        # pdfname = str(employee_dict[name].get('row')) + '.pdf'
        pdfname = ".\\my_folder\\payroll_{}.pdf".format(rownum)
        pdfkit.from_file(htmlfile, pdfname, configuration=config ,options=pdfkit_options)
        file_path = os.path.abspath(pdfname)
        filepathes.append(file_path)
    return filepathes




def main():
    wb = load_workbook(filename="sample.xlsx", data_only=True)
    sheet = wb.active
    emps = employees(sheet)



if __name__ == "__main__":
    main()


