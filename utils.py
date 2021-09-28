# Built-in packages
import smtplib
import os
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from string import Template

# External Packages
import fitz
import pdfkit
from openpyxl import load_workbook

# Project Modules
import config


def sheet(filename):
    wb = load_workbook(filename, data_only=True)
    sh = wb.active
    return sh


def headers(sheet):
    # if we want to skipe last cols if cell.value is not None
    header_cells = sheet[1]
    header = [cell.value for cell in header_cells]
    return header


def employee_data(employee):
    k = [
        "row",  # سطر
        "name",  # اسم کارمند
        "days_worked",  # روز کارد
        "daily_pay",  #  دستمزد روزانه
        "extra_wh",  #  ساعت اضافه کار
        "daily_mp",  #  نرخ ماموریت روزانه
        "extra_whp",  #  نرخ اضافه کار
        "base_pay",  #  پایه حقوق
        "work_pay",  #  مبلغ کارکرد
        "extra_work_pay",  #  مبلغ اضافه کار
        "house_right",  #  حق مسکن
        "extra_help",  #  کمک هزینه اقلام مصرفی
        "child_right",  #   حق بچه
        "commiute",  #  ایاب ذهاب
        "extra_undirect",  #  مزایای غیر مستقیم
        "work_extra",  #  فوق العاده شغل
        "other_benefit",  # سایر مزایا
        "u_payment",  #  حقوق ناخالص
        "seven_ensure",  #  هفت درصد بیمه
        "to_tax",  #  مشمول مالیات
        "tax",  #  مالیات
        "payemt_extra",  # خالص حقوق و مزایا
        "rem",  # علی الحساب
        "loan",  #  وام قرض الحسنه
        "other_insur",  #  سایر بیمه تکمیلی
        "other_req",  # سایر مطالبات
        "mission",  #  ماموریت
        "other",  #  سایر
        "payment",  #  قابل پرداخت
        "name2",  #  اسم
    ]
    v = [cell.value for cell in employee]
    data = dict(zip(k, v))
    employee = {}
    for k, v in data.items():
        if v == None:
            employee[k] = ""
        else:
            employee[k] = v
    # adding extra options
    ew = employee.get("extra_whp")
    employee["year"] = "1399"
    employee["month"] = "مهر"
    employee["extra_whp"] = f"{ew:.2f}"
    return employee


def employees(sheet):
    # skip hedaer
    rows = list(sheet.iter_rows())[2:]
    # if value is not that mean we found an employee
    employees = [row for row in rows if row[0].value is not None]
    return employees[:-1]  # skipe last row


# Create employee dict by employee name
def employee(employees, name=None):
    if name is not None:
        for emp in employees:
            if emp[1].value == name:
                return employee_data(emp)
    return None


# cols: row, name, email
def items(employees, column_name="name"):
    columns = {"row": 0, "name": 1, "email": 30}
    index = columns.get(column_name)
    return [employee[index].value for employee in employees]


def read_template(filename="template.html"):
    with open(filename, "r", encoding="utf-8") as template_file:
        template = Template(template_file.read())
    return template


# create payroll for a employee (get employee dict)
def template_to_html(employee, template, filename="payroll.html"):
    payroll = template.substitute(employee)
    with open(filename, "w", encoding="utf-8") as html_file:
        html_file.write(payroll)
    return filename


def html_to_pdf(html_filename="payroll.html", pdf_filename="payroll.pdf"):
    options = {
        "encoding": "UTF-8",
        "page-size": "A5",
        "margin-top": "0cm",
        "margin-bottom": "0cm",
        "margin-left": "0cm",
        "margin-right": "0cm",
        "dpi": 400,
        "enable-local-file-access": "",
    }
    if os.name == "nt":
        wk_path = pdfkit.configuration(
            wkhtmltopdf=r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
        )
        pdfkit.from_file(
            html_filename, pdf_filename, options=options, configuration=wk_path
        )
        return pdf_filename

    pdfkit.from_file(html_filename, pdf_filename, options=options)
    return pdf_filename


def pdf_to_image(pdf_path, img_name="payroll.jpg"):
    doc = fitz.open(pdf_path)
    page = doc[0]  # payroll is just one page
    pix = page.getPixmap()
    pix.writeImage(img_name)
    return pix


def get_image_bytes(pix):
    pix = fitz.Pixmap(pix, 0) if pix.alpha else pix  # PPM does not support transparency
    imgdata = pix.getImageData("ppm")  # extremely fast!
    return imgdata


def email_connection():
    # # set up the SMTP server
    s = smtplib.SMTP(host=config.EMAIL_HOST, port=config.EMAIL_PORT)
    s.starttls()
    s.login(config.EMAIL_HOST_USER, config.EMAIL_HOST_PASSWORD)
    return s


def send_mail(conn, filename, email, extra):
    if email is None:
        return False

    msg = MIMEMultipart()  # create a message
    name = extra.get("NAME")
    message_template = read_template("message.html")
    # add in the actual person name to the message template
    message = message_template.substitute(extra)

    # setup the parameters of the message
    msg["From"] = config.EMAIL_HOST_USER
    msg["To"] = email
    msg["Subject"] = "شرکت سامان آبراه"

    # add in the message body
    msg.attach(MIMEText(message, "html"))

    # create pdf to send
    sh = sheet(filename)
    emps = employees(sh)
    emp = employee(emps, name)
    template = read_template()
    payroll_html = template_to_html(emp, template)

    filename = html_to_pdf(payroll_html, "payroll.pdf")
    # Open PDF file in binary mode
    with open(filename, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
    encoders.encode_base64(part)
    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )
    msg.attach(part)
    # send the message via the server set up earlier.
    conn.send_message(msg)
    del msg
    return True


if __name__ == "__main__":
    # names = ["saeid"]
    # emails = ["saeidgholami101@gmail.com"]
    conn = email_connection()
    name = "سعید غلامی"
    email = "saeidgholami101.com"
    # result = send_mail(conn, name, email)
    result = ""
    if result:
        print("Sent")
    else:
        print("not send")
