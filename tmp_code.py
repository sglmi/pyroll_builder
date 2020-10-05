from openpyxl import load_workbook
from string import Template
import pdfkit


wb = load_workbook(filename="sample.xlsx", data_only=True)
sheet = wb.active

header_cells = sheet[1]
header = [cell.value.strip() for cell in header_cells if cell.value is not None]

# first employee skip last 6 empty columns
# skip hedaer
rows = list(sheet.iter_rows())[2:]
# extract all employees from sheet(worksheet)

employees = []
for row in rows:
    if row[0].value is not None:  # found an employee
        employees.append(row)

# skip last employee cuse it's not a employee
employees = employees[:-1]

emp1 = sheet[3][:-6]
# dw: روز کارکرد, dp: day paeid,
k = [
    "row",  # سطر
    "name",  # اسم کارمند
    "days_worked",  # روز کارد
    "daily_pay",  #  دستمزد روزانه
    "extra_wh",  #  ساعت اضافه کار
    "daily_mp",  #  نرخ ماموریت روزانه
    "extra_whp",  #  نرخ اضافه کار
    "base_pay",  #  پایه حقوق
    "work_pay",  #  مبلغ کارکرد
    "extra_work_pay",  #  مبلغ اضافه کار
    "house_right",  #  حق مسکن
    "extra_help",  #  کمک هزینه اقلام مصرفی
    "child_right",  #   حق بچه
    "commiute",  #  ایاب ذهاب
    "extra_undirect",  #  مزایای غیر مستقیم
    "work_extra",  #  فوق العاده شغل
    "other_benefit",  # سایر مزایا
    "u_payment",  #  حقوق ناخالص
    "seven_ensure",  #  هفت درصد بیمه
    "to_tax",  #  مشمول مالیات
    "tax",  #  مالیات
    "payemt_extra",  # خالص حقوق و مزایا
    "rem",  # علی الحساب
    "loan",  #  وام قرض الحسنه
    "other_insur",  #  سایر بیمه تکمیلی
    "other_req",  # سایر مطالبات
    "mission",  #  ماموریت
    "other",  #  سایر
    "payment",  #  قابل پرداخت
    "name2",  #  اسم
]

v = [cell.value for cell in emp1]


# an employee with it's data as dict
employee_t = dict(zip(k, v))
employee = {}
for k, v in employee_t.items():
    if v == None:
        employee[k] = ""
    else:
        employee[k] = v

ew = employee.get("extra_whp")


f = open("payroll_template.html", "r")
payroll_template = Template(f.read())
payroll = payroll_template.substitute(
    employee,
    year=1399,
    month="مهر",
    extra_whp=f"{ew:.2f}",
)
f.close()


f = open("payroll.html", "w")
f.write(payroll)
f.close()

options = {
    "encoding": "UTF-8",
    "enable-local-file-access": "",
}

pdfkit.from_file("payroll.html", "payroll.pdf", options=options)
