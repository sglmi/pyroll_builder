# from openpyxl import load_workbook
# from string import Template
# import pdfkit

# wb = load_workbook(filename="sample.xlsx", data_only=True)
# sheet = wb.active

# header_cells = sheet[1]
# # header_cells = sheet[1][:-6]
# header = [cell.value.strip() for cell in header_cells if cell.value is not None]

# # first employee
# emp1 = sheet[3][:-6]
# # dw: روز کارکرد, dp: day paeid,
# k = [
#     "row",  # سطر
#     "name",  # اسم کارمند
#     "days_worked",  # روز کارد
#     "daily_pay",  #  دستمزد روزانه
#     "extra_wh",  #  ساعت اضافه کار
#     "daily_mp",  #  نرخ ماموریت روزانه
#     "extra_whp",  #  نرخ اضافه کار
#     "base_pay",  #  پایه حقوق
#     "work_pay",  #  مبلغ کارکرد
#     "extra_work_pay",  #  مبلغ اضافه کار
#     "house_right",  #  حق مسکن
#     "extra_help",  #  کمک هزینه اقلام مصرفی
#     "child_right",  #   حق بچه
#     "commiute",  #  ایاب ذهاب
#     "extra_undirect",  #  مزایای غیر مستقیم
#     "work_extra",  #  فوق العاده شغل
#     "other_benefit",  # سایر مزایا
#     "u_payment",  #  حقوق ناخالص
#     "seven_ensure",  #  هفت درصد بیمه
#     "to_tax",  #  مشمول مالیات
#     "tax",  #  مالیات
#     "payemt_extra",  # خالص حقوق و مزایا
#     "rem",  # علی الحساب
#     "loan",  #  وام قرض الحسنه
#     "other_insur",  #  سایر بیمه تکمیلی
#     "other_req",  # سایر مطالبات
#     "mission",  #  ماموریت
#     "other",  #  سایر
#     "payment",  #  قابل پرداخت
#     "name2",  #  اسم
# ]

# v = [cell.value for cell in emp1]


# # an employee with it's data as dict
# employee_t = dict(zip(k, v))
# employee = {}
# for k, v in employee_t.items():
#     if v == None:
#         employee[k] = ""
#     else:
#         employee[k] = v

# ew = employee.get("extra_whp")
# f = open("xpayroll_template.html", "r", encoding="utf-8")
# payroll_template = Template(f.read())
# payroll = payroll_template.substitute(
#     employee, year=1399, month="مهر", extra_whp=f"{ew:.2f}"
# )
# f.close()


# f = open("payroll.html", "w", encoding="utf-8")
# f.write(payroll)
# f.close()

# pdfkit_options = {
#     "encoding": "UTF-8",
#     "enable-local-file-access": "",
# }
# # path_wkthmltopdf = b'C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe'
# # config = pdfkit.configuration(wkhtmltopdf=path_wkthmltopdf)


# pdfkit.from_file("payroll.html", "payroll.pdf", options=pdfkit_options)

from openpyxl import load_workbook
from string import Template
import pdfkit

def pull_name(fname):

wb = load_workbook(filename=fname, data_only=True)
sheet = wb.active

header_cells = sheet[1]
# header_cells = sheet[1][:-6]
header = [cell.value.strip() for cell in header_cells if cell.value is not None]

# first employee
a_list = []
k = [
"row", # سطر
"name", # اسم کارمند
"days_worked", # روز کارد
"daily_pay", # دستمزد روزانه
"extra_wh", # ساعت اضافه کار
"daily_mp", # نرخ ماموریت روزانه
"extra_whp", # نرخ اضافه کار
"base_pay", # پایه حقوق
"work_pay", # مبلغ کارکرد
"extra_work_pay", # مبلغ اضافه کار
"house_right", # حق مسکن
"extra_help", # کمک هزینه اقلام مصرفی
"child_right", # حق بچه
"commiute", # ایاب ذهاب
"extra_undirect", # مزایای غیر مستقیم
"work_extra", # فوق العاده شغل
"other_benefit", # سایر مزایا
"u_payment", # حقوق ناخالص
"seven_ensure", # هفت درصد بیمه
"to_tax", # مشمول مالیات
"tax", # مالیات
"payemt_extra", # خالص حقوق و مزایا
"rem", # علی الحساب
"loan", # وام قرض الحسنه
"other_insur", # سایر بیمه تکمیلی
"other_req", # سایر مطالبات
"mission", # ماموریت
"other", # سایر
"payment", # قابل پرداخت
"name2", # اسم
]
for i in range(500):
try:
emp = sheet[i][:-6]
v = [cell.value for cell in emp]
if v[-1] == None:
continue
else:
employee_t = dict(zip(k, v))
if employee_t.get('name') == None:
continue
if employee_t.get('name') == 'نام ونام خانوادگی':
continue

a_list.append(employee_t)
except Exception as e:
pass

#employee_t = dict(zip(k, v))
all_names = []
for i in range(len(a_list)):
all_names.append('{} {}'.format(a_list[i].get('name'), i+1))

# with open('allnames.txt', 'w', encoding='utf-8') as f:
# f.write('\n'.join(all_names))
return all_names

if _name_ == '__main__':
pull_name()