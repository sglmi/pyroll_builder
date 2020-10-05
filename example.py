# from string import Template

# # reading payroll html template
# f = open("payroll_template.html", "r", encoding="utf-8")
# payroll = f.read()
# f.close()

# # create payroll template and substitute it's keys
# payroll_template = Template(payroll)
# payroll = payroll_template.substitute(compay_name="زیرتان قولی")


# ## Create a payroll html file
# f = open("payroll.html", "w")
# f.write(payroll)
# f.close()


# # convert html to pdf
# import pdfkit

# pdfkit.from_file("payroll.html", "payroll.pdf")


from xhtml2pdf import pisa  # import python module

# # Define your data
# f = open("payroll.html")
# source_html = f.read()
# output_filename = "test.pdf"
# f.close()
# import io

# # Utility function
# def convert_html_to_pdf(source_html, output_filename):
#     # open output file for writing (truncated binary)
#     result_file = open(output_filename, "wb")

#     # convert HTML to PDF
#     pisa_status = pisa.CreatePDF(
#         source_html,
#         dest=result_file,
#         encoding="utf-8"
#         # the HTML to convert
#     )  # file handle to recieve result

#     # close output file
#     result_file.close()  # close output file

#     # return False on success and True on errors
#     return pisa_status.err


# # Main program
# if __name__ == "__main__":
#     pisa.showLogging()
#     convert_html_to_pdf(source_html, output_filename)

from openpyxl import load_workbook
from string import Template
import pdfkit

wb = load_workbook(filename="sample.xlsx", data_only=True)
sheet = wb.active

header_cells = sheet[1]
header = [cell.value.strip() for cell in header_cells if cell.value is not None]

# print(header)
# first employee skip last 6 empty columns
employees = []
emp1 = sheet[3][:-6]
# print(sheet)

for row in sheet.iter_rows():
    print(row[3].value)